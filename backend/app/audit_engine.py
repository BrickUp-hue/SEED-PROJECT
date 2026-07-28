"""
Motor de auditoría de trazabilidad SEED (RGC Coffee).
Procesa 1 archivo maestro de productores + N archivos ET (por contenedor/embarque),
y produce 3 tablas de hallazgos:
  1. traceability_ranges: rango de fechas de compra por contenedor vs shipment month
  2. producer_volume: volumen comprado vs producción anual declarada, por productor y año calendario
  3. producers_not_in_seed: compras a productores que no existen en la base maestra SEED

Diseñado para escalar a muchos archivos ET sin re-leer todo cada vez: cada ET se
procesa una sola vez y se guarda un registro de transacciones intermedio (parquet/csv)
que luego se agrega.
"""
import openpyxl
import datetime
import re
from pathlib import Path

# ---------- Utilidades ----------

def parse_any_date(value):
    """Convierte a datetime.date manejando datetime nativo, números de serie de Excel
    (cuando la celda no tiene formato de fecha aplicado) o strings dd/mm/yyyy."""
    if value is None or value == '':
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        # Número de serie de Excel (sistema 1900, con el bug del año bisiesto 1900)
        if 20000 <= value <= 60000:  # rango razonable para fechas ~1954-2064
            try:
                return (datetime.datetime(1899, 12, 30) + datetime.timedelta(days=value)).date()
            except (OverflowError, ValueError):
                return None
        return None
    if isinstance(value, str):
        v = value.strip()
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y', '%d-%m-%Y'):
            try:
                return datetime.datetime.strptime(v, fmt).date()
            except ValueError:
                continue
        # último recurso: buscar patrón dd/mm/yyyy embebido
        m = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', v)
        if m:
            d, mo, y = map(int, m.groups())
            try:
                return datetime.date(y, mo, d)
            except ValueError:
                return None
    return None


def months_between(d1, d2):
    """Diferencia en meses (fraccional aprox.) entre dos fechas: d2 - d1."""
    return (d2.year - d1.year) * 12 + (d2.month - d1.month) + (d2.day - d1.day) / 30.0


def normalize_farm_id(value):
    """Normaliza un Farm ID para el cruce: quita espacios, pasa a mayúsculas.
    Blindaje ante variaciones de formato entre archivos (mayúsc/minúsc, espacios extra)."""
    if value is None:
        return None
    return str(value).strip().upper()


def parse_purchase_or_delivery_date(value):
    """
    Parser específico para las columnas de fecha de COMPRA (K) y ENTREGA (N) del
    Suppliers Purchases List — confirmado empíricamente que el 100% de las fechas
    nativas de Excel en estas columnas están invertidas (día/mes intercambiados):
    de 5,587+ fechas nativas revisadas, el 100% tenía día<=12, lo cual es
    estadísticamente imposible para fechas reales (se esperaría ~39%). Esto ocurre
    porque el archivo se llenó en formato DD/MM pero Excel las auto-convirtió
    usando MM/DD (configuración regional), guardando el valor ya invertido.

    Las fechas en TEXTO (no nativas) NO tienen este problema — Excel no pudo
    auto-convertirlas (por eso quedaron como texto) y ya se parsean correctamente
    con parse_any_date (proban DD/MM y MM/DD, usando la única interpretación válida).

    Shipment Month y Contract Date NO usan este parser — esas celdas usan formato
    con el mes escrito (ej. "13-Jan-2026"), sin ambigüedad posicional, y no están
    afectadas (confirmado: sus días llegan hasta 31, no solo <=12).
    """
    if isinstance(value, datetime.datetime):
        try:
            return datetime.date(value.year, value.day, value.month)  # swap día<->mes
        except ValueError:
            return value.date()  # si el swap no da fecha válida, se deja como está (no debería pasar)
    if isinstance(value, datetime.date):
        try:
            return datetime.date(value.year, value.day, value.month)
        except ValueError:
            return value
    # Strings y números de serie: parse_any_date ya los maneja correctamente
    return parse_any_date(value)


# ---------- Carga del maestro de productores ----------

def load_master(master_path):
    """
    Devuelve:
      - active: dict farm_id -> {"ann_prod_60kg_bags": float, "entity_name": str}
      - retirados: set de farm_id retirados
    """
    wb = openpyxl.load_workbook(master_path, data_only=True)

    active = {}
    ws = wb['Update SEED']
    headers = {cell.value: cell.column for cell in ws[1] if cell.value}
    col_farm_id = headers.get('Farm ID')
    col_prod = headers.get('ANN. PROD. 60 KG BAGS')
    col_entity = headers.get('Entity Name')
    for row in ws.iter_rows(min_row=2):
        farm_id = row[col_farm_id - 1].value
        if not farm_id:
            continue
        farm_id = normalize_farm_id(farm_id)
        prod = row[col_prod - 1].value if col_prod else None
        entity = row[col_entity - 1].value if col_entity else None
        active[farm_id] = {
            "ann_prod_60kg_bags": prod,
            "entity_name": entity,
        }

    retirados = set()
    ws2 = wb['Retirados']
    for row in ws2.iter_rows(min_row=2, min_col=1, max_col=1):
        v = row[0].value
        if v:
            retirados.add(normalize_farm_id(v))

    return active, retirados


# ---------- Carga de un archivo ET ----------

def find_header_row(ws, expected_col_b_label=None, search_range=10):
    """Busca la fila de encabezados reales (algunos ET tienen filas de título arriba)."""
    for r in range(1, search_range + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 20)]
        non_empty = sum(1 for v in row_vals if v not in (None, ''))
        if non_empty >= 5:
            return r
    return 1


def load_et_file(et_path, correct_suspicious_shipment_year_to=None):
    """
    Devuelve un dict con:
      - container_id (nombre de archivo, ya que no siempre hay un ID limpio único)
      - shipment_month: date
      - po_number, seller_name, contract_date (contexto)
      - purchases: lista de dicts {farm_id, farm_name, purchase_date, delivery_date, qty_kg}

    Si correct_suspicious_shipment_year_to se especifica (ej. 2026) y se detecta que
    contract_date > shipment_month (error de digitación confirmado por el usuario),
    se reemplaza el AÑO del shipment_month por ese valor (CORRECCIÓN EXPLÍCITA,
    instruida por el usuario, no un supuesto silencioso).
    """
    wb = openpyxl.load_workbook(et_path, data_only=True)
    fname = Path(et_path).stem

    ws1 = wb['SELLERS SALE + SHIPMENT DETAILS']
    shipment_month = parse_any_date(ws1['B5'].value)
    po_number = ws1['B3'].value
    contract_date = parse_any_date(ws1['B4'].value)
    seller_name = ws1['B7'].value

    shipment_month_corrected = False
    shipment_month_original = shipment_month
    if (correct_suspicious_shipment_year_to and contract_date and shipment_month
            and contract_date > shipment_month):
        try:
            shipment_month = shipment_month.replace(year=correct_suspicious_shipment_year_to)
            shipment_month_corrected = True
        except ValueError:
            pass

    # Localizar la hoja de "Suppliers Purchases List" (nombre puede variar levemente)
    sheet_name = None
    for name in wb.sheetnames:
        if 'SUPPLIER' in name.upper() and 'PURCHASE' in name.upper():
            sheet_name = name
            break
    if sheet_name is None:
        raise ValueError(f"No se encontró hoja de compras a proveedores en {fname}")

    ws = wb[sheet_name]
    header_row = find_header_row(ws)

    purchases = []
    for r in range(header_row + 1, ws.max_row + 1):
        farm_id = ws.cell(row=r, column=9).value    # I
        farm_name = ws.cell(row=r, column=10).value  # J
        purchase_date_raw = ws.cell(row=r, column=11).value  # K
        delivery_date_raw = ws.cell(row=r, column=14).value  # N
        qty_kg = ws.cell(row=r, column=16).value     # P

        if farm_id is None and qty_kg is None:
            continue

        purchases.append({
            "farm_id": normalize_farm_id(farm_id) if farm_id else None,
            "farm_name": farm_name,
            "purchase_date": parse_purchase_or_delivery_date(purchase_date_raw),
            "purchase_date_raw": purchase_date_raw,
            "delivery_date": parse_purchase_or_delivery_date(delivery_date_raw),
            "qty_kg": qty_kg if isinstance(qty_kg, (int, float)) else None,
        })

    return {
        "container_id": fname,
        "po_number": po_number,
        "shipment_month": shipment_month,
        "shipment_month_original": shipment_month_original,
        "shipment_month_corrected": shipment_month_corrected,
        "contract_date": contract_date,
        "seller_name": seller_name,
        "purchases": purchases,
    }


# ---------- Análisis 1: rangos de trazabilidad por contenedor ----------

def analyze_traceability_ranges(et_data, min_months=3, max_months=4):
    """
    Devuelve un dict con:
      - resumen del contenedor (rango de fechas, % fuera de rango, etc.)
      - 'out_of_range_by_month': lista de {year_month, n_transacciones, kg_total, fecha_mas_lejana}
      - 'after_shipment': lista de transacciones cuya fecha de compra es POSTERIOR al shipment month (hallazgo crítico)
    """
    purchases_with_date = [p for p in et_data["purchases"] if p["purchase_date"]]
    dates = [p["purchase_date"] for p in purchases_with_date]
    n_total = len(et_data["purchases"])
    n_missing_date = sum(1 for p in et_data["purchases"] if p["purchase_date"] is None)
    shipment_month = et_data["shipment_month"]

    if not dates:
        return {
            "container_id": et_data["container_id"],
            "po_number": et_data["po_number"],
            "shipment_month": shipment_month,
            "contract_date": et_data.get("contract_date"),
            "shipment_month_suspicious": bool(et_data.get("contract_date") and shipment_month and et_data.get("contract_date") > shipment_month),
            "date_min": None, "date_max": None,
            "n_purchases": n_total, "n_missing_date": n_missing_date,
            "n_out_of_range": None,
            "pct_out_of_range": None,
            "flag": "SIN FECHAS VÁLIDAS",
            "out_of_range_by_month": [],
            "after_shipment": [],
        }

    date_min, date_max = min(dates), max(dates)

    # Hallazgo de calidad de dato (confirmado por el usuario como error de digitación en la fuente):
    # si la fecha de contrato es posterior al shipment month declarado, el shipment month
    # de ESTE archivo es sospechoso. Se deja como flag informativo, no se corrige el shipment month.
    contract_date = et_data.get("contract_date")
    shipment_month_suspicious = bool(contract_date and shipment_month and contract_date > shipment_month)

    out_of_range_purchases = []
    after_shipment_purchases = []

    if shipment_month:
        for p in purchases_with_date:
            d = p["purchase_date"]
            m = months_between(d, shipment_month)  # meses que la compra antecede al embarque

            # Hallazgo crítico: compra fechada DESPUÉS del shipment month
            if d > shipment_month:
                after_shipment_purchases.append(p)
                continue  # ya está fuera de rango por definición, no duplicar abajo

            # Fuera de la ventana esperada de 3-4 meses antes (regla estricta, sin margen)
            if not (min_months <= m <= max_months):
                out_of_range_purchases.append(p)

    n_out = len(out_of_range_purchases) + len(after_shipment_purchases)
    pct_out = round(100 * n_out / len(dates), 1) if dates else None

    # SUPUESTO EXPLÍCITO (no confirmado): de las fuera-de-rango, ¿cuántas caerían
    # DENTRO de una tolerancia de 5 meses? Podría deberse a un acuerdo particular con SBUX.
    # Esto NO se resta del conteo oficial de fuera-de-rango; se muestra aparte como supuesto.
    within_5mo_assumption = [p for p in out_of_range_purchases if min_months <= months_between(p["purchase_date"], shipment_month) <= 5]
    n_within_5mo_assumption = len(within_5mo_assumption)
    kg_within_5mo_assumption = round(sum(p["qty_kg"] or 0 for p in within_5mo_assumption), 1)

    flag = "OK"
    if pct_out and pct_out > 30:
        flag = "⚠️ ALTO % FUERA DE RANGO"
    elif pct_out and pct_out > 0:
        flag = "REVISAR"
    if after_shipment_purchases:
        flag = "🔴 CRÍTICO: COMPRAS DESPUÉS DEL EMBARQUE"

    # Agrupar fuera-de-rango (que NO son "after shipment") por mes calendario
    from collections import defaultdict
    by_month = defaultdict(lambda: {"n": 0, "kg": 0.0, "farthest_date": None})
    for p in out_of_range_purchases:
        d = p["purchase_date"]
        ym = f"{d.year}-{d.month:02d}"
        bucket = by_month[ym]
        bucket["n"] += 1
        bucket["kg"] += p["qty_kg"] or 0
        if bucket["farthest_date"] is None or months_between(d, shipment_month) > months_between(bucket["farthest_date"], shipment_month):
            bucket["farthest_date"] = d

    out_of_range_by_month = [
        {
            "year_month": ym,
            "n_transacciones": v["n"],
            "kg_total": round(v["kg"], 1),
            "fecha_mas_lejana": v["farthest_date"],
            "meses_antes_del_embarque": round(months_between(v["farthest_date"], shipment_month), 1) if shipment_month else None,
        }
        for ym, v in sorted(by_month.items())
    ]

    after_shipment_summary = [
        {
            "farm_id": p["farm_id"],
            "farm_name": p["farm_name"],
            "purchase_date": p["purchase_date"],
            "qty_kg": p["qty_kg"],
            "dias_despues_del_embarque": (p["purchase_date"] - shipment_month).days if shipment_month else None,
        }
        for p in after_shipment_purchases
    ]

    return {
        "container_id": et_data["container_id"],
        "po_number": et_data["po_number"],
        "shipment_month": shipment_month,
        "shipment_month_original": et_data.get("shipment_month_original"),
        "shipment_month_corrected": et_data.get("shipment_month_corrected", False),
        "contract_date": contract_date,
        "shipment_month_suspicious": shipment_month_suspicious,
        "date_min": date_min, "date_max": date_max,
        "n_purchases": n_total, "n_missing_date": n_missing_date,
        "n_out_of_range": n_out,
        "pct_out_of_range": pct_out,
        "flag": flag,
        "out_of_range_by_month": out_of_range_by_month,
        "after_shipment": after_shipment_summary,
        # --- SUPUESTO NO CONFIRMADO (mostrar aparte, no fusionar con el resultado oficial) ---
        "supuesto_tolerancia_5_meses": {
            "nota": "SUPUESTO: podría existir un acuerdo particular con SBUX que amplíe la ventana a 5 meses. No confirmado — de las transacciones fuera de rango (regla estricta 3-4 meses), estas son las que SÍ calificarían si se aplicara una tolerancia de 5 meses.",
            "n_transacciones": n_within_5mo_assumption,
            "kg_total": kg_within_5mo_assumption,
        },
        "scenarios": compute_tolerance_scenarios(purchases_with_date, shipment_month, min_months),
    }


def compute_tolerance_scenarios(purchases_with_date, shipment_month, min_months=3, scenario_max_months=(4, 5, 6)):
    """
    SUPUESTO / SENSIBILIDAD (no oficial): para cada escenario de tolerancia máxima
    (ej. 4, 5, 6 meses antes del embarque, manteniendo el mínimo de 3 meses),
    cuenta cuántas transacciones (con fecha de compra <= shipment_month, es decir
    excluyendo las 'después del embarque' que son siempre críticas sin importar tolerancia)
    quedarían FUERA de rango bajo cada supuesto.
    """
    if not shipment_month:
        return {m: None for m in scenario_max_months}

    # Solo transacciones con compra <= shipment_month (las posteriores son críticas aparte, siempre)
    eligible = [p for p in purchases_with_date if p["purchase_date"] and p["purchase_date"] <= shipment_month]
    n_eligible = len(eligible)

    results = {}
    for max_m in scenario_max_months:
        n_out = 0
        kg_out = 0.0
        for p in eligible:
            m = months_between(p["purchase_date"], shipment_month)
            if not (min_months <= m <= max_m):
                n_out += 1
                kg_out += p["qty_kg"] or 0
        results[max_m] = {
            "n_out_of_range": n_out,
            "pct_out_of_range": round(100 * n_out / n_eligible, 1) if n_eligible else None,
            "kg_out_of_range": round(kg_out, 1),
        }
    return results


# ---------- Análisis 2 y 3: volumen por productor + productores no-SEED ----------

def analyze_purchases(et_data, active_producers, retirados):
    """
    Devuelve:
      - producer_rows: lista de transacciones enriquecidas (para agregación posterior)
      - not_in_seed_rows: compras a productores que no están ni en activos ni en retirados
    """
    producer_rows = []
    not_in_seed_rows = []

    for p in et_data["purchases"]:
        farm_id = p["farm_id"]
        if not farm_id:
            continue

        qty = p["qty_kg"] or 0
        purchase_date = p["purchase_date"]
        year = purchase_date.year if purchase_date else None

        in_active = farm_id in active_producers
        in_retirados = farm_id in retirados

        status = "ACTIVO" if in_active else ("RETIRADO" if in_retirados else "NO ENCONTRADO EN SEED")

        producer_rows.append({
            "container_id": et_data["container_id"],
            "farm_id": farm_id,
            "farm_name": p["farm_name"],
            "purchase_date": purchase_date,
            "purchase_year": year,
            "qty_kg": qty,
            "seed_status": status,
        })

        if status == "NO ENCONTRADO EN SEED":
            not_in_seed_rows.append({
                "container_id": et_data["container_id"],
                "farm_id": farm_id,
                "farm_name": p["farm_name"],
                "purchase_date": purchase_date,
                "qty_kg": qty,
            })

    return producer_rows, not_in_seed_rows


def aggregate_producer_volume(all_producer_rows, active_producers):
    """
    Agrega por (farm_id, purchase_year) la suma de qty_kg comprado,
    y calcula % vs producción anual declarada (convertida de sacos 60kg a kg).
    """
    from collections import defaultdict
    agg = defaultdict(float)
    names = {}

    for row in all_producer_rows:
        if row["purchase_year"] is None:
            continue
        key = (row["farm_id"], row["purchase_year"])
        agg[key] += row["qty_kg"] or 0
        names[row["farm_id"]] = row["farm_name"]

    results = []
    for (farm_id, year), total_kg in agg.items():
        info = active_producers.get(farm_id)
        prod_bags_60kg = info["ann_prod_60kg_bags"] if info else None
        prod_kg = (prod_bags_60kg * 60) if isinstance(prod_bags_60kg, (int, float)) else None

        pct = (total_kg / prod_kg * 100) if prod_kg and prod_kg > 0 else None

        flag = "SIN DATO DE PRODUCCIÓN"
        if pct is not None:
            if pct > 100:
                flag = "🔴 SUPERA 100% DE PRODUCCIÓN"
            elif pct < 20:
                flag = "🟡 MENOS DEL 20% VENDIDO"
            elif pct < 50:
                flag = "🟠 ENTRE 20% Y 50% (bajo el mínimo esperado)"
            else:
                flag = "✅ CUMPLE (≥50%)"

        results.append({
            "farm_id": farm_id,
            "farm_name": info["entity_name"] if info else names.get(farm_id),
            "year": year,
            "total_purchased_kg": round(total_kg, 1),
            "declared_prod_60kg_bags": prod_bags_60kg,
            "declared_prod_kg": round(prod_kg, 1) if prod_kg else None,
            "pct_of_declared": round(pct, 1) if pct is not None else None,
            "flag": flag,
        })

    return results
